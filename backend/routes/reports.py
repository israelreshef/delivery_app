from flask import Blueprint, request, jsonify, Response, make_response
from models import db, Delivery, Invoice, User, Courier, Expense, CompanySettings
from utils.decorators import token_required, role_required
import logging
from datetime import datetime, timedelta
import csv
import io
from sqlalchemy import func, and_
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/revenue', methods=['GET'])
@token_required
@role_required(['admin', 'finance_admin'])
def get_revenue_report(current_user):
    """דוח הכנסות לפי טווח תאריכים"""
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        # Default to this month
        if not start_date_str:
            start_date = datetime.utcnow().replace(day=1)
        else:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            
        if not end_date_str:
            end_date = datetime.utcnow()
        else:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) # Include the end day

        # Query Revenue grouped by day
        revenue_data = db.session.query(
            func.date(Invoice.issue_date).label('date'),
            func.sum(Invoice.total_amount).label('total'),
            func.count(Invoice.id).label('count')
        ).filter(
            Invoice.issue_date >= start_date,
            Invoice.issue_date < end_date,
            Invoice.status.in_(['paid', 'sent']) # sent is also revenue theoretically
        ).group_by(
            func.date(Invoice.issue_date)
        ).all()
        
        result = []
        total_period_revenue = 0
        
        for r in revenue_data:
            amount = float(r.total)
            result.append({
                'date': r.date.strftime('%Y-%m-%d'),
                'amount': amount,
                'count': r.count
            })
            total_period_revenue += amount
            
        return jsonify({
            'period': {'start': start_date.strftime('%Y-%m-%d'), 'end': (end_date - timedelta(days=1)).strftime('%Y-%m-%d')},
            'daily_breakdown': result,
            'total_revenue': total_period_revenue
        }), 200
        
    except Exception as e:
        logging.error(f"Error generating revenue report: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/export', methods=['GET'])
@token_required
@role_required(['admin', 'finance_admin'])
def export_csv(current_user):
    """ייצוא דוחות ל-CSV"""
    try:
        report_type = request.args.get('type', 'orders')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Date range required'}), 400
            
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
        
        si = io.StringIO()
        cw = csv.writer(si)
        
        filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.csv"
        
        if report_type == 'orders':
            # Export Orders
            cw.writerow(['Order ID', 'Date', 'Status', 'Customer', 'Courier', 'Pickup', 'Dropoff', 'Price', 'Distance'])
            
            orders = Delivery.query.filter(
                Delivery.created_at >= start_date,
                Delivery.created_at < end_date
            ).all()
            
            for o in orders:
                cw.writerow([
                    o.order_number,
                    o.created_at.strftime('%Y-%m-%d %H:%M'),
                    o.status,
                    o.customer.full_name if o.customer else 'Guest',
                    o.courier.full_name if o.courier else 'Unassigned',
                    o.pickup_point.address.city if o.pickup_point else '',
                    o.delivery_point.address.city if o.delivery_point else '',
                    o.invoice.total_amount if o.invoice else 0,
                    o.distance_km or 0
                ])
                
        elif report_type == 'revenue':
            # Export Revenue
            cw.writerow(['Invoice ID', 'Date', 'Customer', 'Amount', 'VAT', 'Total', 'Status'])
            
            invoices = Invoice.query.filter(
                Invoice.issue_date >= start_date,
                Invoice.issue_date < end_date
            ).all()
            
            for i in invoices:
                cw.writerow([
                    i.invoice_number,
                    i.issue_date.strftime('%Y-%m-%d'),
                    i.customer.full_name if i.customer else 'Unknown',
                    i.subtotal,
                    i.vat_amount,
                    i.total_amount,
                    i.status
                ])
        
        else:
            return jsonify({'error': 'Invalid report type'}), 400
            
        output = make_response(si.getvalue().encode('utf-8-sig')) # utf-8-sig for Hebrew Excel support
        output.headers["Content-Disposition"] = f"attachment; filename={filename}"
        output.headers["Content-type"] = "text/csv"
        return output
        
    except Exception as e:
        logging.error(f"Error exporting CSV: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/regulatory', methods=['GET'])
@token_required
@role_required(['admin', 'finance_admin'])
def export_regulatory_excel(current_user):
    """ייצוא קובץ אקסל מקיף לדוחות רגולטוריים (מע"מ, מס הכנסה, וכו')"""
    try:
        from models import Expense  # Import here to avoid circular dependencies if needed
        
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if not start_date_str or not end_date_str:
            # Default to current month
            today = datetime.utcnow()
            start_date = today.replace(day=1)
            # end_date is first of next month
            if today.month == 12:
                end_date = datetime(today.year + 1, 1, 1)
            else:
                end_date = datetime(today.year, today.month + 1, 1)
        else:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)

        # Create Excel Workbook
        wb = openpyxl.Workbook()
        
        # ---------------------------------------------------------
        # 1. דוח רווח והפסד (P&L) - Sheet 1 (Active)
        # ---------------------------------------------------------
        ws_pl = wb.active
        ws_pl.title = "רווח והפסד"
        ws_pl.sheet_view.rightToLeft = True
        
        # Calculate Revenues
        invoices = Invoice.query.filter(
            Invoice.issue_date >= start_date,
            Invoice.issue_date < end_date,
            Invoice.status.in_(['paid', 'sent'])
        ).all()
        
        total_revenue_before_vat = sum(float(i.subtotal) for i in invoices)
        
        # Calculate Expenses
        expenses = Expense.query.filter(
            Expense.expense_date >= start_date,
            Expense.expense_date < end_date
        ).all()
        
        total_expenses_before_vat = sum(float(e.amount) for e in expenses)
        
        # P&L Layout
        ws_pl.append(["דוח רווח והפסד", f"לתקופה שבין {start_date.strftime('%d/%m/%Y')} ל-{(end_date-timedelta(days=1)).strftime('%d/%m/%Y')}"])
        ws_pl.append([])
        ws_pl.append(["הכנסות (ללא מע\"מ)", total_revenue_before_vat])
        ws_pl.append(["הוצאות (ללא מע\"מ)", total_expenses_before_vat])
        ws_pl.append([])
        ws_pl.append(["רווח נקי (לפני מס)", total_revenue_before_vat - total_expenses_before_vat])
        
        # Styling Pl
        for row in ws_pl.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")
        
        # ---------------------------------------------------------
        # 2. דוח מע"מ (VAT) - Sheet 2
        # ---------------------------------------------------------
        ws_vat = wb.create_sheet(title="דו\"ח מע\"מ")
        ws_vat.sheet_view.rightToLeft = True
        
        total_vat_collected = sum(float(i.vat_amount) for i in invoices)
        total_vat_paid = sum(float(e.vat_amount) for e in expenses)
        vat_to_pay = total_vat_collected - total_vat_paid
        
        ws_vat.append(["דוח מע\"מ תקופתי"])
        ws_vat.append([])
        ws_vat.append(["מע\"מ עסקאות (שנגבה מלקוחות)", total_vat_collected])
        ws_vat.append(["מע\"מ תשומות (ששולם לספקים)", total_vat_paid])
        ws_vat.append([])
        ws_vat.append(["מע\"מ לתשלום (להחזר אם שלילי)", vat_to_pay])
        
        # Detail Tables
        ws_vat.append([])
        ws_vat.append(["פירוט עסקאות:"])
        ws_vat.append(["מספר חשבונית", "תאריך", "לקוח", "סכום לפני מע\"מ", "מע\"מ", "סה\"כ"])
        for i in invoices:
            ws_vat.append([
                i.invoice_number, 
                i.issue_date.strftime('%d/%m/%Y'), 
                i.customer.full_name if i.customer else 'Unknown',
                float(i.subtotal), float(i.vat_amount), float(i.total_amount)
            ])
            
        ws_vat.append([])
        ws_vat.append(["פירוט תשומות:"])
        ws_vat.append(["מזהה הוצאה", "תאריך", "תיאור", "קטגוריה", "סכום לפני מע\"מ", "מע\"מ", "סה\"כ"])
        for e in expenses:
            ws_vat.append([
                e.id, 
                e.expense_date.strftime('%d/%m/%Y'), 
                e.description, e.category,
                float(e.amount), float(e.vat_amount), float(e.total_amount)
            ])

        # ---------------------------------------------------------
        # 3. דוח מקדמות מס הכנסה (Income Tax Advances) - Sheet 3
        # ---------------------------------------------------------
        ws_tax = wb.create_sheet(title="מקדמות מס הכנסה")
        ws_tax.sheet_view.rightToLeft = True
        
        # Hardcoded advance rate for example - should be adjustable in settings
        ADVANCE_RATE = 0.05 # 5%
        tax_advance_to_pay = total_revenue_before_vat * ADVANCE_RATE
        
        ws_tax.append(["חישוב מקדמות מס הכנסה"])
        ws_tax.append([])
        ws_tax.append(["מחזור עסקאות (ללא מע\"מ)", total_revenue_before_vat])
        ws_tax.append(["אחוז מקדמה", f"{ADVANCE_RATE * 100}%"])
        ws_tax.append([])
        ws_tax.append(["סכום מקדמה לתשלום", tax_advance_to_pay])

        # ---------------------------------------------------------
        # 4. ביטוח לאומי (National Insurance) - Sheet 4
        # ---------------------------------------------------------
        ws_ni = wb.create_sheet(title="ביטוח לאומי")
        ws_ni.sheet_view.rightToLeft = True
        
        # Bituach Leumi is typically calculated on the NET PROFIT, not gross revenue.
        net_profit = total_revenue_before_vat - total_expenses_before_vat
        # Rough estimation for self-employed (usually a tiered percentage)
        NI_RATE = 0.09 # Rough average estimation ~9%
        ni_to_pay = net_profit * NI_RATE if net_profit > 0 else 0
        
        ws_ni.append(["הערכת תשלומי ביטוח לאומי לעצמאי"])
        ws_ni.append([])
        ws_ni.append(["רווח נקי", net_profit])
        ws_ni.append(["הערכת אחוז ב\"ל", f"{NI_RATE * 100}%"])
        ws_ni.append([])
        ws_ni.append(["סכום מוערך לתשלום", ni_to_pay])
        
        # Format columns width for all sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = 20

        # Save to memory stream
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        filename = f"regulatory_report_{start_date.strftime('%Y%m')}.xlsx"
        
        response = make_response(out.read())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
        
    except Exception as e:
        logging.error(f"Error generating regulatory Excel report: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/vat-pcn874', methods=['GET'])
@token_required
@role_required(['admin', 'finance_admin'])
def export_pcn874(current_user):
    """
    ייצוא קובץ PCN 874 לדיווח מע"מ דיגיטלי (פורמט טקסט תקני).
    עבור פשטות המימוש הראשוני, נפיק גרסת Excel מפורטת שעוקבת אחרי המבנה הנדרש.
    """
    try:
        month = request.args.get('month', datetime.utcnow().month, type=int)
        year = request.args.get('year', datetime.utcnow().year, type=int)
        
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)

        settings = CompanySettings.query.first() or CompanySettings(legal_name="TZIR Logistics", tax_id="000000000")

        # 1. Revenues (Sales)
        invoices = Invoice.query.filter(
            Invoice.issue_date >= start_date,
            Invoice.issue_date < end_date,
            Invoice.status.in_(['paid', 'sent'])
        ).all()

        # 2. Expenses (Purchases)
        expenses = Expense.query.filter(
            Expense.expense_date >= start_date.date(),
            Expense.expense_date < end_date.date()
        ).all()

        si = io.StringIO()
        cw = csv.writer(si)
        
        cw.writerow(['קובץ דיווח מע"מ PCN 874 - טיוטה'])
        cw.writerow(['חברה:', settings.legal_name, 'ח"פ:', settings.tax_id])
        cw.writerow(['תקופה:', f"{month}/{year}"])
        cw.writerow([])
        cw.writerow(['סוג רשומה', 'מספר מסמך', 'תאריך', 'ח"פ/ת.ז ספק/לקוח', 'סכום לפני מע"מ', 'מע"מ', 'קוד מע"מ'])
        
        # Sales records
        for inv in invoices:
            cw.writerow(['עסקאות', inv.invoice_number, inv.issue_date.strftime('%d/%m/%Y'), inv.customer.tax_id if inv.customer and hasattr(inv.customer, 'tax_id') else '', float(inv.subtotal), float(inv.vat_amount), '1'])
            
        # Purchase records
        for exp in expenses:
            cw.writerow(['תשומות', exp.id, exp.expense_date.strftime('%d/%m/%Y'), exp.courier.tax_id if exp.courier else exp.vendor_name, float(exp.base_amount), float(exp.vat_amount), '1'])

        output = make_response(si.getvalue().encode('utf-8-sig'))
        output.headers["Content-Disposition"] = f"attachment; filename=PCN874_{year}_{month}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/contractors-856', methods=['GET'])
@token_required
@role_required(['admin', 'finance_admin'])
def export_contractors_856(current_user):
    """
    דו"ח 856: ריכוז תשלומים וניכוי מס במקור לקבלנים (שליחים).
    """
    try:
        year = request.args.get('year', datetime.utcnow().year, type=int)
        
        expenses = Expense.query.filter(
            Expense.is_contractor_invoice == True,
            func.extract('year', Expense.expense_date) == year
        ).all()

        # Group by courier
        summary = {}
        for exp in expenses:
            if not exp.courier_id: continue
            cid = exp.courier_id
            if cid not in summary:
                summary[cid] = {
                    'name': exp.courier.full_name,
                    'tax_id': exp.courier.tax_id,
                    'gross': 0,
                    'withholding': 0,
                    'net': 0
                }
            summary[cid]['gross'] += float(exp.base_amount + exp.vat_amount)
            summary[cid]['withholding'] += float(exp.withholding_tax_deducted)
            summary[cid]['net'] += float(exp.total_amount)

        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['דו"ח 856 - ריכוז תשלומים וניכוי מס במקור לקבלנים', f'שנת {year}'])
        cw.writerow(['שם השליח', 'ח"פ/ת.ז', 'תשלום ברוטו', 'ניכוי מס במקור', 'תשלום נטו'])
        
        for cid, data in summary.items():
            cw.writerow([data['name'], data['tax_id'], data['gross'], data['withholding'], data['net']])

        output = make_response(si.getvalue().encode('utf-8-sig'))
        output.headers["Content-Disposition"] = f"attachment; filename=Report856_{year}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/financial-overview', methods=['GET'])
@token_required
@role_required(['admin', 'finance_admin'])
def get_financial_summary(current_user):
    """מרכז נתונים פיננסיים לעמוד הראשי של הניהול"""
    try:
        month = request.args.get('month', datetime.utcnow().month, type=int)
        year = request.args.get('year', datetime.utcnow().year, type=int)
        
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)

        # Revenue
        rev_stats = db.session.query(
            func.sum(Invoice.subtotal).label('net'),
            func.sum(Invoice.vat_amount).label('vat'),
            func.sum(Invoice.total_amount).label('gross')
        ).filter(Invoice.issue_date >= start_date, Invoice.issue_date < end_date).first()

        # Expenses
        exp_stats = db.session.query(
            func.sum(Expense.base_amount).label('net'),
            func.sum(Expense.vat_amount).label('vat'),
            func.sum(Expense.withholding_tax_deducted).label('withholding'),
            func.sum(Expense.total_amount).label('gross')
        ).filter(Expense.expense_date >= start_date.date(), Expense.expense_date < end_date.date()).first()

        return jsonify({
            'period': f"{month}/{year}",
            'revenue': {
                'net': float(rev_stats.net or 0),
                'vat': float(rev_stats.vat or 0),
                'gross': float(rev_stats.gross or 0)
            },
            'expenses': {
                'net': float(exp_stats.net or 0),
                'vat': float(exp_stats.vat or 0),
                'withholding': float(exp_stats.withholding or 0),
                'gross': float(exp_stats.gross or 0)
            },
            'profit_loss': float((rev_stats.net or 0) - (exp_stats.net or 0))
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
